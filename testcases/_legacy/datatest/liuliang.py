import requests
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill
import os
import warnings
from urllib3.exceptions import InsecureRequestWarning

# 禁用SSL警告
warnings.filterwarnings('ignore', category=InsecureRequestWarning)

# 配置参数
API_URL = "https://api.agora.io/beta/realtime/traffic"
APP_ID = "edca9ba1d5c945d2a031370981a58892"
CHANNEL_NAME = "1181509"
AUTH_TOKEN = "agora token=Basic NTk0Njc5MmIxMTg4NDkzN2IzNDk2MGU3YWU5MjQ4Mzc6YWQzZjM1MjNmYjYwNDMwNmEzZDBkMDM0OWRhYjgxMWQ="  # 替换为你的实际鉴权Token


def get_timestamps():
    """获取当天9点的时间戳和当前时间戳"""
    today = datetime.now().date()
    nine_am = datetime.combine(today, datetime.strptime("09:00:00", "%H:%M:%S").time())
    end_time = datetime.now()
    return int(nine_am.timestamp()), int(end_time.timestamp())


def call_agora_api():
    """调用Agora API获取流量数据"""
    start_time, end_time = get_timestamps()

    headers = {
        "Authorization": AUTH_TOKEN,
        "Content-Type": "application/json",
        "Accept": "application/json",
        "User-Agent": "AgoraTrafficMonitor/1.0"
    }

    params = {
        "channel_name": CHANNEL_NAME,
        "appid": APP_ID,
        "start_time": start_time,
        "end_time": end_time
    }

    try:
        response = requests.get(
            API_URL,
            headers=headers,
            params=params,
            verify=False,
            timeout=15
        )
        response.raise_for_status()
        return response.json()
    except requests.exceptions.RequestException as e:
        print(f"API请求失败: {str(e)}")
        return None


def convert_units(traffic_kb, interval_seconds=300, binary=True):
    """
    单位转换计算
    :param traffic_kb: 下行流量(KB)
    :param interval_seconds: 时间间隔(秒)
    :param binary: 是否使用1024进制
    :return: 转换后的所有单位
    """
    # KB → MB 转换
    divisor = 1024 if binary else 1000
    traffic_mb = traffic_kb / divisor

    # KB → KB/s 转换
    traffic_kbps = traffic_kb / interval_seconds

    # KB/s → Mbps 转换
    traffic_mbps = (traffic_kbps * 8) / 1000  # 1 KB/s = 0.008 Mbps

    return {
        "mb": traffic_mb,
        "kbps": traffic_kbps,
        "mbps": traffic_mbps,
        "binary_standard": binary
    }


def process_data(api_data, binary_units=True):
    """处理API返回数据并执行单位转换"""
    if not api_data or api_data.get("code") != 200:
        print(f"无效的API响应: {api_data}")
        return None

    traffic_list = api_data.get("data", {}).get("traffic_data_list", [])
    if not traffic_list:
        print("流量数据为空")
        return None

    processed_data = []
    for i, item in enumerate(traffic_list):
        ts = item.get("ts", 0)
        traffic_kb = item.get("down_traffic", 0)

        # 计算实际时间间隔
        interval = 300  # 默认5分钟间隔
        if i > 0:
            prev_ts = traffic_list[i - 1].get("ts", ts - interval)
            interval = max(ts - prev_ts, 60)  # 最小60秒

        # 单位转换
        converted = convert_units(traffic_kb, interval, binary_units)

        processed_data.append({
            "timestamp": ts,
            "time": datetime.fromtimestamp(ts).strftime("%Y-%m-%d %H:%M:%S"),
            "original_kb": traffic_kb,
            "interval_seconds": interval,
            **converted
        })

    return processed_data


def save_to_excel(data, filename="traffic_report.xlsx"):
    """保存数据到Excel文件"""
    if not data:
        print("无有效数据可保存")
        return

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "流量分析报告"

        # 写入表头
        headers = [
            "时间戳", "时间", "时间间隔(s)",
            "原始数据(KB)", "MB(" + ("1024" if data[0]["binary_standard"] else "1000") + "进制)",
            "速率(KB/s)", "速率(Mbps)"
        ]
        ws.append(headers)

        # 设置表头样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        # 写入数据
        for item in data:
            ws.append([
                item["timestamp"],
                item["time"],
                item["interval_seconds"],
                item["original_kb"],
                item["mb"],
                item["kbps"],
                item["mbps"]
            ])

        # 设置数字格式
        number_format = "0.000"
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            row[4].number_format = number_format  # MB
            row[5].number_format = number_format  # KB/s
            row[6].number_format = number_format  # Mbps

        # 自动调整列宽
        for col in ws.columns:
            max_length = max(len(str(cell.value)) for cell in col)
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[col[0].column_letter].width = adjusted_width

        # 保存文件
        wb.save(filename)
        print(f"报告已保存: {os.path.abspath(filename)}")

    except Exception as e:
        print(f"保存Excel文件时出错: {str(e)}")


def main():
    print("=== Agora流量分析工具 ===")
    print(f"频道: {CHANNEL_NAME}")
    print(f"应用ID: {APP_ID}")

    # 获取API数据
    print("\n正在获取流量数据...")
    api_data = call_agora_api()

    if api_data:
        # 处理数据（默认使用1024进制）
        processed_data = process_data(api_data, binary_units=True)

        if processed_data:
            # 保存Excel报告
            save_to_excel(processed_data)

    print("\n分析完成！")


if __name__ == "__main__":
    main()
